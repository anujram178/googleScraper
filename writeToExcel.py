import openpyxl
from pathlib import Path
import search 
import sys

# setting path
# S = Search()
helper = search.Search()
xlsx_file = Path(sys.argv[1])
print(sys.argv[1])
print(xlsx_file)

# read file
wb_obj = openpyxl.load_workbook(xlsx_file) 

# read active sheet from workbook
wsheet = wb_obj.active



print(wsheet.max_row, wsheet.max_column)

rowNumber = 0
for row in wsheet.iter_rows(max_row=wsheet.max_row):
	rowNumber += 1
	if rowNumber != 1:
		nameCell = 'B' + str(rowNumber)
		stateCell = 'G' + str(rowNumber)
		googleQuery = wsheet[nameCell].value.lower() + ' ' +  wsheet[stateCell].value.lower()
		results = helper.scrapeResults(googleQuery)
		urlDict = helper.findUrl(results)
		webAddressCell = 'J' + str(rowNumber)
		if urlDict['exists'] == True:
			wsheet[webAddressCell] = urlDict['link']
			print("this is the link: ", urlDict['link'])
		wb_obj.save(xlsx_file)
